from __future__ import annotations

import io
from datetime import date, datetime, timezone

import openpyxl

from app.collectors.bea_iea_support import parse_bea_iip_article
from app.collectors.public_data import (
    CollectedSeries,
    PublicDataCollector,
    YahooBar,
    _active_contracts_for_day,
    _build_bea_iip_histories,
    _build_foreign_duration_sponsorship_history,
    _build_iea_oil_security_histories,
    _build_basis_proxy_series,
    _build_consumer_credit_composite_history,
    _build_front_contract_history,
    _build_period_change_history,
    _build_payroll_tax_base_history,
    _build_pointwise_series,
    _build_ranked_stress_history,
    _build_synthetic_ctd_basis_history,
    _build_treasury_depth_history,
    _compute_auction_stress_histories,
    _compute_auction_stress_history,
    _densify_daily,
    _generate_contract_specs,
    _parse_nyfed_household_credit_workbook,
)


def test_densify_daily_carries_last_observation_forward() -> None:
    history = _densify_daily(
        [(date(2026, 3, 10), 1.0), (date(2026, 3, 12), 3.0)],
        start=date(2026, 3, 10),
        end=date(2026, 3, 13),
    )

    assert [point[1] for point in history] == [1.0, 1.0, 3.0, 3.0]


def test_build_pointwise_series_uses_latest_available_inputs() -> None:
    history = _build_pointwise_series(
        start=date(2026, 3, 10),
        end=date(2026, 3, 12),
        inputs={
            'SOFR': [(date(2026, 3, 10), 4.50)],
            'DFF': [(date(2026, 3, 10), 4.20), (date(2026, 3, 12), 4.25)],
        },
        formula=lambda values: round((values['SOFR'] - values['DFF']) * 100.0, 2),
    )

    assert [point[1] for point in history] == [30.0, 30.0, 25.0]


def test_compute_auction_stress_history_generates_daily_step_series() -> None:
    rows = [
        {
            'auction_date': '2026-03-10',
            'security_type': 'Note',
            'security_term': '10-Year',
            'bid_to_cover_ratio': '2.70',
            'high_yield': '4.55',
            'offering_amt': '39000000000',
            'indirect_bidder_accepted': '25000000000',
        },
        {
            'auction_date': '2026-03-12',
            'security_type': 'Bond',
            'security_term': '30-Year',
            'bid_to_cover_ratio': '2.25',
            'high_yield': '4.88',
            'offering_amt': '22000000000',
            'indirect_bidder_accepted': '12000000000',
        },
        {
            'auction_date': '2026-03-25',
            'security_type': 'Note',
            'security_term': '5-Year',
            'bid_to_cover_ratio': '2.20',
            'high_yield': '4.45',
            'offering_amt': '70000000000',
            'indirect_bidder_accepted': '39000000000',
            'primary_dealer_accepted': '15000000000',
        },
    ]

    history = _compute_auction_stress_history(rows, start=date(2026, 3, 10), end=date(2026, 3, 26))

    assert len(history) == 17
    assert history[0][1] < history[-1][1]
    assert history[-2][1] == history[-1][1]


def test_auction_stress_history_uses_trailing_distribution_without_future_lookahead() -> None:
    rows = [
        {
            'auction_date': '2026-01-10',
            'security_type': 'Note',
            'security_term': '10-Year',
            'bid_to_cover_ratio': '2.70',
            'high_yield': '4.20',
            'offering_amt': '39000000000',
            'indirect_bidder_accepted': '25000000000',
            'primary_dealer_accepted': '9000000000',
        },
        {
            'auction_date': '2026-02-10',
            'security_type': 'Note',
            'security_term': '10-Year',
            'bid_to_cover_ratio': '2.60',
            'high_yield': '4.25',
            'offering_amt': '40000000000',
            'indirect_bidder_accepted': '24000000000',
            'primary_dealer_accepted': '10000000000',
        },
        {
            'auction_date': '2026-03-10',
            'security_type': 'Note',
            'security_term': '10-Year',
            'bid_to_cover_ratio': '2.10',
            'high_yield': '4.60',
            'offering_amt': '39000000000',
            'indirect_bidder_accepted': '18000000000',
            'primary_dealer_accepted': '14000000000',
        },
    ]

    histories = _compute_auction_stress_histories(rows, start=date(2026, 1, 10), end=date(2026, 3, 12))
    clearing_history = histories['auction_clearing_stress']

    jan10 = next(value for timestamp, value in clearing_history if timestamp.date() == date(2026, 1, 10))
    feb10 = next(value for timestamp, value in clearing_history if timestamp.date() == date(2026, 2, 10))
    mar10 = next(value for timestamp, value in clearing_history if timestamp.date() == date(2026, 3, 10))

    assert jan10 == 50.0
    assert feb10 > jan10
    assert mar10 > feb10


def test_auction_histories_capture_belly_and_cluster_stress_from_repeated_5y_weakness() -> None:
    rows = [
        {
            'auction_date': '2026-01-27',
            'security_type': 'Note',
            'security_term': '5-Year',
            'bid_to_cover_ratio': '2.55',
            'high_yield': '4.18',
            'offering_amt': '70000000000',
            'indirect_bidder_accepted': '43000000000',
            'primary_dealer_accepted': '12000000000',
        },
        {
            'auction_date': '2026-02-25',
            'security_type': 'Note',
            'security_term': '5-Year',
            'bid_to_cover_ratio': '2.32',
            'high_yield': '4.31',
            'offering_amt': '70000000000',
            'indirect_bidder_accepted': '43600000000',
            'primary_dealer_accepted': '8950000000',
        },
        {
            'auction_date': '2026-03-26',
            'security_type': 'Note',
            'security_term': '5-Year',
            'bid_to_cover_ratio': '2.15',
            'high_yield': '4.44',
            'offering_amt': '70000000000',
            'indirect_bidder_accepted': '39800000000',
            'primary_dealer_accepted': '16500000000',
        },
        {
            'auction_date': '2026-03-10',
            'security_type': 'Note',
            'security_term': '10-Year',
            'bid_to_cover_ratio': '2.52',
            'high_yield': '4.35',
            'offering_amt': '39000000000',
            'indirect_bidder_accepted': '24000000000',
            'primary_dealer_accepted': '10500000000',
        },
    ]

    histories = _compute_auction_stress_histories(rows, start=date(2026, 1, 27), end=date(2026, 3, 27))

    belly = histories['auction_belly_clearing_stress']
    cluster = histories['auction_coupon_cluster_stress']
    composite = histories['auction_stress']

    jan = next(value for timestamp, value in belly if timestamp.date() == date(2026, 1, 27))
    feb = next(value for timestamp, value in belly if timestamp.date() == date(2026, 2, 25))
    mar = next(value for timestamp, value in belly if timestamp.date() == date(2026, 3, 26))
    cluster_mar = next(value for timestamp, value in cluster if timestamp.date() == date(2026, 3, 26))
    composite_mar = next(value for timestamp, value in composite if timestamp.date() == date(2026, 3, 26))

    assert feb > jan
    assert mar > feb
    assert cluster_mar > 0
    assert composite_mar >= mar * 0.5


def test_basis_proxy_series_becomes_more_negative_with_wider_rate_gap_and_fx_vol() -> None:
    fx = [
        (date(2026, 3, 10), 1.10),
        (date(2026, 3, 11), 1.11),
        (date(2026, 3, 12), 1.08),
        (date(2026, 3, 13), 1.13),
        (date(2026, 3, 14), 1.09),
    ]
    usd_rate = [(date(2026, 3, 10), 4.4), (date(2026, 3, 14), 4.5)]
    local_rate = [(date(2026, 3, 10), 1.5), (date(2026, 3, 14), 1.0)]

    history = _build_basis_proxy_series(
        start=date(2026, 3, 10),
        end=date(2026, 3, 14),
        fx_observations=fx,
        usd_rate_observations=usd_rate,
        local_rate_observations=local_rate,
        rate_multiplier=6.5,
        vol_multiplier=2.5,
        floor_value=5.0,
        ceiling_value=80.0,
        vol_window=3,
    )

    assert history[0][1] < 0
    assert history[-1][1] < history[0][1]


def test_active_contract_selection_rolls_to_next_month() -> None:
    specs = _generate_contract_specs('CL', '.NYM', start=date(2026, 1, 1), end=date(2026, 3, 31), monthly=True, roll_lead_days=35)
    selected = _active_contracts_for_day(specs, date(2026, 3, 5))
    assert selected[0].symbol.endswith('.NYM')
    assert selected[0].delivery_month <= selected[1].delivery_month


def test_front_contract_history_and_treasury_metrics_build() -> None:
    specs = _generate_contract_specs('ZN', '.CBT', start=date(2026, 3, 1), end=date(2026, 3, 5), monthly=False, roll_lead_days=20)
    first, second = _active_contracts_for_day(specs, date(2026, 3, 1))
    bar_map = {
        first.symbol: {
            date(2026, 3, 1): YahooBar(timestamp=None, close=111.0, high=111.2, low=110.8, volume=100000.0),
            date(2026, 3, 2): YahooBar(timestamp=None, close=110.8, high=111.0, low=110.4, volume=120000.0),
            date(2026, 3, 3): YahooBar(timestamp=None, close=110.6, high=111.1, low=110.3, volume=90000.0),
            date(2026, 3, 4): YahooBar(timestamp=None, close=110.9, high=111.3, low=110.5, volume=95000.0),
            date(2026, 3, 5): YahooBar(timestamp=None, close=111.1, high=111.4, low=110.7, volume=98000.0),
        },
        second.symbol: {
            date(2026, 3, 1): YahooBar(timestamp=None, close=110.7, high=110.9, low=110.5, volume=80000.0),
            date(2026, 3, 2): YahooBar(timestamp=None, close=110.6, high=110.8, low=110.4, volume=81000.0),
            date(2026, 3, 3): YahooBar(timestamp=None, close=110.4, high=110.7, low=110.2, volume=82000.0),
            date(2026, 3, 4): YahooBar(timestamp=None, close=110.6, high=110.9, low=110.3, volume=83000.0),
            date(2026, 3, 5): YahooBar(timestamp=None, close=110.8, high=111.0, low=110.5, volume=84000.0),
        },
    }
    front_history = _build_front_contract_history(specs, bar_map, start=date(2026, 3, 1), end=date(2026, 3, 5))
    depth = _build_treasury_depth_history(front_history, window=3)
    seven_year = [(bar.timestamp, 4.05 + index * 0.02) for index, bar in enumerate(front_history)]
    ten_year = [(bar.timestamp, 4.20 + index * 0.02) for index, bar in enumerate(front_history)]
    basis = _build_synthetic_ctd_basis_history(front_history, seven_year, ten_year, window=3)

    assert len(front_history) == 5
    assert len(depth) == 5
    assert len(basis) == 5
    assert all(value >= 0 for _, value in depth)
    assert all(value >= 0 for _, value in basis)


def test_synthetic_ctd_basis_history_rises_with_larger_futures_cash_yield_gap() -> None:
    front_history = [
        YahooBar(timestamp=datetime(2026, 3, day, tzinfo=timezone.utc), close=110.9 + offset, high=111.0, low=110.5, volume=90000.0)
        for offset, day in enumerate(range(1, 6))
    ]
    seven_year = [
        (datetime(2026, 3, day, tzinfo=timezone.utc), 4.00 + 0.01 * index)
        for index, day in enumerate(range(1, 6))
    ]
    ten_year = [
        (datetime(2026, 3, day, tzinfo=timezone.utc), 4.20 + 0.01 * index)
        for index, day in enumerate(range(1, 6))
    ]

    calm = _build_synthetic_ctd_basis_history(front_history, seven_year, ten_year, window=3)
    stressed_front_history = [
        YahooBar(timestamp=bar.timestamp, close=bar.close + 2.5, high=bar.high + 2.5, low=bar.low + 2.5, volume=bar.volume)
        for bar in front_history
    ]
    stressed = _build_synthetic_ctd_basis_history(stressed_front_history, seven_year, ten_year, window=3)

    assert len(calm) == len(stressed) == 5
    assert stressed[-1][1] > calm[-1][1]



def test_period_change_history_supports_average_change_and_yoy_percent() -> None:
    observations = [
        (date(2025 + ((month - 1) // 12), ((month - 1) % 12) + 1, 1), 100.0 + month * 10.0)
        for month in range(1, 16)
    ]

    average_change = _build_period_change_history(observations, start=date(2026, 1, 1), end=date(2026, 3, 31), periods_back=3, pct=False, divisor=3.0)
    yoy_change = _build_period_change_history(observations, start=date(2026, 1, 1), end=date(2026, 3, 31), periods_back=12, pct=True)

    assert round(average_change[0][1], 2) == 10.0
    assert round(yoy_change[0][1], 2) == 109.09


def test_payroll_tax_base_history_combines_payrolls_wages_and_hours() -> None:
    payrolls = [(date(2025 + ((month - 1) // 12), ((month - 1) % 12) + 1, 1), 1000.0 + month * 10.0) for month in range(1, 16)]
    wages = [(date(2025 + ((month - 1) // 12), ((month - 1) % 12) + 1, 1), 20.0 + month * 0.1) for month in range(1, 16)]
    hours = [(date(2025 + ((month - 1) // 12), ((month - 1) % 12) + 1, 1), 34.0 + month * 0.05) for month in range(1, 16)]

    history = _build_payroll_tax_base_history(payrolls, wages, hours, start=date(2026, 1, 1), end=date(2026, 3, 31))

    assert len(history) >= 1
    assert history[0][1] > 0


def test_ranked_stress_history_rises_with_higher_levels_and_worsening_momentum() -> None:
    observations = [
        (date(2025, 3, 31), 2.1),
        (date(2025, 6, 30), 2.3),
        (date(2025, 9, 30), 2.6),
        (date(2025, 12, 31), 3.1),
        (date(2026, 3, 31), 4.0),
    ]

    history = _build_ranked_stress_history(observations, start=date(2025, 3, 31), end=date(2026, 3, 31), window=5)

    assert history[0][1] < history[-1][1]
    assert history[-1][1] >= 80.0


def test_parse_nyfed_household_credit_workbook_reads_card_and_auto_all_age_series() -> None:
    workbook = openpyxl.Workbook()
    default_sheet = workbook.active
    workbook.remove(default_sheet)
    card_sheet = workbook.create_sheet('Page 27 Data')
    auto_sheet = workbook.create_sheet('Page 26 Data')
    for sheet in (card_sheet, auto_sheet):
        sheet.append(['title'])
        sheet.append(['Percent'])
        sheet.append(['quarter', '18-29', '30-39', '40-49', '50-59', '60-69', '70+', 'all'])
    card_sheet.append(['25:Q3', None, None, None, None, None, None, 4.9])
    card_sheet.append(['25:Q4', None, None, None, None, None, None, 5.4])
    auto_sheet.append(['25:Q3', None, None, None, None, None, None, 2.2])
    auto_sheet.append(['25:Q4', None, None, None, None, None, None, 2.6])
    buffer = io.BytesIO()
    workbook.save(buffer)
    workbook.close()

    card_history, auto_history = _parse_nyfed_household_credit_workbook(
        buffer.getvalue(),
        card_sheet='Page 27 Data',
        auto_sheet='Page 26 Data',
    )

    assert card_history[-1] == (date(2025, 12, 31), 5.4)
    assert auto_history[-1] == (date(2025, 12, 31), 2.6)


def test_consumer_credit_composite_requires_broad_confirmation_to_break_out() -> None:
    start = date(2026, 3, 1)
    end = date(2026, 3, 3)

    single_block = _build_consumer_credit_composite_history(
        [
            ('bank_quality', [(datetime(2026, 3, 1, tzinfo=timezone.utc), 88.0)], 0.30),
            ('household_ccp_quality', [(datetime(2026, 3, 1, tzinfo=timezone.utc), 35.0)], 0.20),
            ('lender_tightening', [(datetime(2026, 3, 1, tzinfo=timezone.utc), 30.0)], 0.20),
            ('borrowing_cost', [(datetime(2026, 3, 1, tzinfo=timezone.utc), 28.0)], 0.15),
            ('household_strain', [(datetime(2026, 3, 1, tzinfo=timezone.utc), 32.0)], 0.15),
        ],
        start=start,
        end=end,
    )
    broad_breakout = _build_consumer_credit_composite_history(
        [
            ('bank_quality', [(datetime(2026, 3, 1, tzinfo=timezone.utc), 82.0)], 0.30),
            ('household_ccp_quality', [(datetime(2026, 3, 1, tzinfo=timezone.utc), 76.0)], 0.20),
            ('lender_tightening', [(datetime(2026, 3, 1, tzinfo=timezone.utc), 74.0)], 0.20),
            ('borrowing_cost', [(datetime(2026, 3, 1, tzinfo=timezone.utc), 70.0)], 0.15),
            ('household_strain', [(datetime(2026, 3, 1, tzinfo=timezone.utc), 66.0)], 0.15),
        ],
        start=start,
        end=end,
    )

    assert single_block[-1][1] <= 62.0
    assert broad_breakout[-1][1] > single_block[-1][1]


def test_collect_consumer_credit_series_builds_composite_from_fred_nyfed_and_household_inputs(monkeypatch) -> None:
    collector = PublicDataCollector()
    start = date(2026, 1, 1)
    end = date(2026, 3, 31)

    fred_map = {
        'DRCCLACBS': [(date(2025, 3, 31), 2.0), (date(2025, 6, 30), 2.2), (date(2025, 9, 30), 2.5), (date(2025, 12, 31), 2.9), (date(2026, 3, 31), 3.4)],
        'DRALACBS': [(date(2025, 3, 31), 1.6), (date(2025, 6, 30), 1.7), (date(2025, 9, 30), 1.9), (date(2025, 12, 31), 2.1), (date(2026, 3, 31), 2.4)],
        'CORCCACBS': [(date(2025, 3, 31), 3.0), (date(2025, 6, 30), 3.2), (date(2025, 9, 30), 3.4), (date(2025, 12, 31), 3.8), (date(2026, 3, 31), 4.4)],
        'CORCACBS': [(date(2025, 3, 31), 2.0), (date(2025, 6, 30), 2.1), (date(2025, 9, 30), 2.2), (date(2025, 12, 31), 2.4), (date(2026, 3, 31), 2.7)],
        'TERMCBCCALLNS': [(date(2025, 1, 1), 20.1), (date(2025, 6, 1), 20.7), (date(2025, 12, 1), 21.5), (date(2026, 3, 1), 22.2)],
        'DRTSCLCC': [(date(2025, 3, 31), 8.0), (date(2025, 6, 30), 12.0), (date(2025, 9, 30), 18.0), (date(2025, 12, 31), 28.0), (date(2026, 3, 31), 35.0)],
        'STDSOTHCONS': [(date(2025, 3, 31), 4.0), (date(2025, 6, 30), 7.0), (date(2025, 9, 30), 12.0), (date(2025, 12, 31), 18.0), (date(2026, 3, 31), 24.0)],
    }

    def fake_fetch_fred_csv(self, client, series_id, fetch_start, fetch_end, transform=None):
        observations = fred_map[series_id]
        if transform is None:
            return observations
        return [(observed_at, transform(value)) for observed_at, value in observations]

    def fake_fetch_nyfed_household_credit_transition_histories(self, client):
        return (
            [(date(2025, 3, 31), 4.3), (date(2025, 6, 30), 4.5), (date(2025, 9, 30), 4.9), (date(2025, 12, 31), 5.2), (date(2026, 3, 31), 5.6)],
            [(date(2025, 3, 31), 2.0), (date(2025, 6, 30), 2.1), (date(2025, 9, 30), 2.2), (date(2025, 12, 31), 2.4), (date(2026, 3, 31), 2.7)],
        )

    monkeypatch.setattr(PublicDataCollector, '_fetch_fred_csv', fake_fetch_fred_csv)
    monkeypatch.setattr(PublicDataCollector, '_fetch_nyfed_household_credit_transition_histories', fake_fetch_nyfed_household_credit_transition_histories)

    temp_help = _densify_daily(
        [(date(2025, 12, 31), -1.0), (date(2026, 3, 31), -5.5)],
        start=start,
        end=end,
    )
    tax_base = _densify_daily(
        [(date(2025, 12, 31), 3.2), (date(2026, 3, 31), 0.5)],
        start=start,
        end=end,
    )
    income_squeeze = _densify_daily(
        [(date(2025, 12, 31), 54.0), (date(2026, 3, 31), 70.0)],
        start=start,
        end=end,
    )
    collected = {
        'temp_help_stress': CollectedSeries('temp_help_stress', 'fred/TEMPHELPS-YOY', temp_help),
        'employment_tax_base_proxy': CollectedSeries('employment_tax_base_proxy', 'proxy/fred-payroll-tax-base', tax_base),
        'household_real_income_squeeze': CollectedSeries('household_real_income_squeeze', 'support/income-energy-squeeze', income_squeeze),
    }

    series, status = collector._collect_consumer_credit_series(object(), start, end, collected)

    assert 'consumer_credit_stress' in series
    assert series['consumer_credit_stress'].source == 'support/fred-nyfed-consumer-credit-composite'
    assert series['consumer_credit_stress'].history[-1][1] >= 58.0
    assert 'New York Fed household transitions' in status


def test_parse_bea_iip_article_captures_signed_financial_transactions() -> None:
    raw_html = """
    <html>
      <head>
        <title>SCB, A Look at the U.S. International Investment Position: Fourth Quarter and Year 2025, January 2026</title>
      </head>
      <body>
        <p>The net international investment position decreased from -$26.23 trillion at the end of the third quarter to -$27.54 trillion at the end of the fourth quarter.</p>
        <p>U.S. assets increased by $1.72 trillion to a total of $38.49 trillion.</p>
        <p>U.S. liabilities increased by $3.03 trillion to a total of $66.03 trillion.</p>
        <p>Financial transactions reduced U.S. liabilities by $182.4 billion in the fourth quarter.</p>
      </body>
    </html>
    """

    observation = parse_bea_iip_article(raw_html)

    assert observation is not None
    assert observation.observed_at == date(2025, 12, 31)
    assert observation.net_iip_trillion == -27.54
    assert observation.liability_financial_transactions_billion == -182.4


def test_iea_oil_security_histories_build_cover_and_buffer_stress() -> None:
    start = date(2026, 1, 1)
    end = date(2026, 3, 31)
    observations = [
        type('Obs', (), {'observed_at': date(2025, 12, 31), 'country_name': 'Total IEA net importers', 'total_days': 141.0, 'industry_days': 79.0, 'public_days': 62.0})(),
        type('Obs', (), {'observed_at': date(2025, 12, 31), 'country_name': 'Total IEA Europe', 'total_days': 130.0, 'industry_days': 75.0, 'public_days': 55.0})(),
        type('Obs', (), {'observed_at': date(2025, 12, 31), 'country_name': 'Japan', 'total_days': 208.0, 'industry_days': 91.0, 'public_days': 117.0})(),
        type('Obs', (), {'observed_at': date(2025, 12, 31), 'country_name': 'Korea', 'total_days': 200.0, 'industry_days': 92.0, 'public_days': 108.0})(),
        type('Obs', (), {'observed_at': date(2026, 3, 31), 'country_name': 'Total IEA net importers', 'total_days': 118.0, 'industry_days': 76.0, 'public_days': 42.0})(),
        type('Obs', (), {'observed_at': date(2026, 3, 31), 'country_name': 'Total IEA Europe', 'total_days': 110.0, 'industry_days': 70.0, 'public_days': 40.0})(),
        type('Obs', (), {'observed_at': date(2026, 3, 31), 'country_name': 'Japan', 'total_days': 165.0, 'industry_days': 86.0, 'public_days': 79.0})(),
        type('Obs', (), {'observed_at': date(2026, 3, 31), 'country_name': 'Korea', 'total_days': 158.0, 'industry_days': 84.0, 'public_days': 74.0})(),
    ]

    histories = _build_iea_oil_security_histories(observations, start, end)

    assert histories['iea_oil_cover_days'][-1][1] == 118.0
    assert round(histories['iea_public_stock_share'][-1][1], 2) == 35.59
    assert histories['oil_buffer_depletion_stress'][-1][1] > histories['oil_buffer_depletion_stress'][0][1]
    assert histories['iea_importer_oil_cover_stress'][-1][1] > 50.0


def test_bea_and_foreign_duration_histories_build_structural_sponsorship_overlay() -> None:
    start = date(2026, 1, 1)
    end = date(2026, 3, 31)
    observations = [
        type('Obs', (), {'observed_at': date(2025, 12, 31), 'net_iip_trillion': -27.54, 'liability_financial_transactions_billion': -182.4})(),
        type('Obs', (), {'observed_at': date(2026, 3, 31), 'net_iip_trillion': -28.10, 'liability_financial_transactions_billion': 90.0})(),
    ]

    bea_histories = _build_bea_iip_histories(observations, start, end)
    foreign_duration = _build_foreign_duration_sponsorship_history(
        start,
        end,
        _densify_daily([(date(2025, 12, 31), 56.0), (date(2026, 3, 31), 74.0)], start, end),
        _densify_daily([(date(2025, 12, 31), 52.0), (date(2026, 3, 31), 70.0)], start, end),
        _densify_daily([(date(2025, 12, 31), 8.0), (date(2026, 3, 31), 28.0)], start, end),
        _densify_daily([(date(2025, 12, 31), 48.0), (date(2026, 3, 31), 67.0)], start, end),
        bea_histories['bea_net_iip_burden'],
        bea_histories['bea_foreign_financing_support'],
    )

    assert bea_histories['bea_net_iip_burden'][-1][1] == 28.1
    assert bea_histories['bea_foreign_financing_support'][-1][1] == 90.0
    assert foreign_duration[-1][1] > foreign_duration[0][1]
